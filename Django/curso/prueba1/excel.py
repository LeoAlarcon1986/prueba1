import re
import time
import random
import pytest
#import sys
import openpyxl  # libreria de excel
from playwright.sync_api import Page, expect, Playwright, sync_playwright
from funciones import funciones_globales
from confi_test import set_up , set_up_val_num, set_up_val_pass, set_up_excel

#pytest confi_test.py -s -v
#pytest sesion3.py -s -v --browser-channel=chrome -n 4
#playwright codegen https://www.saucedemo.com/v1/
#pytest -s -v Parametrizar.py -n 3 --html=reportes1.html --self-contained-html --capture=tee-sys
#pytest Parametrizar.py -s -v -n2 --template=html1/index.html --report=reporte_tres.html
#https://pypi.org/project/pytest-html/
#https://pypi.org/project/pytest-reporter-html1/
#https://www.mockaroo.com/
tiempo=0.5
ruta="Proyecto_2/Django/"  
rutaexcel="C:/Proyecto_2/Django/curso/prueba1/datos1.xlsx"
pdf1="C:/Proyecto_2/Django/curso/prueba1/pdf/Pensión Juan Agosto 2023.pdf"
registros=5

archivo=openpyxl.load_workbook(rutaexcel)
def numero_fila(hoja):
    ac=archivo[hoja]
    return ac.max_row

def dato_columna(hoja,fila,col):
    ac=archivo[hoja]
    col=ac.cell(int(fila),int(col))
    return col.value

print(numero_fila("Hoja1"))
print(dato_columna("Hoja1",4,4))

def test_excel(set_up_excel)-> None:
    page=set_up_excel
    f=funciones_globales(page)
    #f.validar_texto("Datos Personales | TestingQaRvn")
    f.scroll(0,600,3)
    print("cargando excel")
    f.esperar(3)
    
    for n in range(2,registros):
      nombre= dato_columna("Hoja1",n,1)
      print(nombre)
      ap= dato_columna("Hoja1",n,2)
      print(ap)
      email= dato_columna("Hoja1",n,3)
      print(email)
      telefono= dato_columna("Hoja1",n,4)
      print(telefono)
      direccion= dato_columna("Hoja1",n,5)
      print(direccion)
        
      f.Texto("//input[contains(@id,'wsf-1-field-21')]",nombre)
      f.Texto("//input[contains(@id,'wsf-1-field-22')]",ap)
      f.Texto("//input[contains(@id,'wsf-1-field-23')]",email)
      f.Texto("//input[contains(@id,'wsf-1-field-24')]",str(telefono))
      f.Texto("//textarea[contains(@id,'wsf-1-field-28')]",direccion)
      f.click("//button[contains(@id,'wsf-1-field-27')]")
      f.esperar(2)
      page.goto("https://testingqarvn.com.es/datos-personales/")
      f.esperar(1)
        
